"""
API de Generador de Certificados
Versión 3.0 - Optimizada para producción
"""
from flask import Flask, request, jsonify, render_template, session, redirect, send_file
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_cors import CORS
import logging
from pythonjsonlogger import jsonlogger
from datetime import datetime
from pydantic import ValidationError
import os

# Módulos propios
from config import config
from database import Database
from cloudinary_storage import CloudinaryStorage
from generator import CertificateGenerator
from schemas import (
    GenerarCertificadosRequest,
    ParticipanteSchema
)
from security import (
    require_admin_ip,
    validate_slug,
    validate_cloudinary_url,
    sanitize_error_message
)

# Configuración de logging estructurado
def setup_logging():
    """Configura logging estructurado con JSON"""
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # Handler para consola
    logHandler = logging.StreamHandler()
    formatter = jsonlogger.JsonFormatter(
        '%(asctime)s %(name)s %(levelname)s %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    logHandler.setFormatter(formatter)
    logger.addHandler(logHandler)

    return logging.getLogger(__name__)


# Setup
logger = setup_logging()
app = Flask(__name__)
app.secret_key = config.SECRET_KEY

# CORS
CORS(app)

# Rate Limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[config.RATELIMIT_DEFAULT] if config.RATELIMIT_ENABLED else []
)
try:
    if config.RATELIMIT_ENABLED and config.RATELIMIT_STORAGE_URL:
        limiter.storage_uri = config.RATELIMIT_STORAGE_URL
        logger.info("Rate limiting habilitado con Redis")
    else:
        logger.warning("Rate limiting usando memoria (no recomendado en producción)")
except Exception as e:
    logger.warning(f"Rate limiting configurado sin Redis: {e}")

# Inicializar servicios (con fallbacks para permitir arranque)
db = None
cloudinary_storage = None
generator = None

try:
    db = Database(config.DATABASE_URL)
    logger.info("Base de datos inicializada")
except Exception as e:
    logger.warning(f"Base de datos no disponible: {e}")
    logger.warning("La app arrancará pero funcionalidad limitada")

try:
    if config.CLOUDINARY_CLOUD_NAME and config.CLOUDINARY_API_KEY:
        cloudinary_storage = CloudinaryStorage(
            cloud_name=config.CLOUDINARY_CLOUD_NAME,
            api_key=config.CLOUDINARY_API_KEY,
            api_secret=config.CLOUDINARY_API_SECRET,
            folder=config.CLOUDINARY_FOLDER
        )
        logger.info("Cloudinary inicializado")
    else:
        logger.warning("Cloudinary no configurado (configurar variables de entorno)")
except Exception as e:
    logger.warning(f"Error al inicializar Cloudinary: {e}")

try:
    if db and os.path.exists(config.TEMPLATE_PATH):
        generator = CertificateGenerator(
            template_path=config.TEMPLATE_PATH,
            database=db,
            cloudinary_storage=cloudinary_storage
        )
        logger.info("Generador de certificados inicializado")
    else:
        logger.warning("Generador no inicializado (requiere DB y template)")
except Exception as e:
    logger.warning(f"Error al inicializar generador: {e}")


# === HEALTH & INFO ===

@app.route('/', methods=['GET'])
def index():
    """Información de la API"""
    total = 0
    if db:
        try:
            total = db.contar_certificados()
        except:
            total = 0

    return jsonify({
        'nombre': config.APP_NAME,
        'version': '3.0',
        'certificados_generados': total,
        'cloudinary_configurado': cloudinary_storage.configured if cloudinary_storage else False,
        'database_configurado': db is not None,
        'endpoints': {
            'health': '/health',
            'generar': '/generar-certificados (POST)',
            'ver_certificado': '/certificado/<slug> (GET)',
            'preview': '/preview/<slug> (GET)',
            'descargar': '/descargar/<slug> (GET)',
            'listar': '/listar-certificados (GET)',
            'buscar_email': '/buscar/email/<email> (GET)',
            'buscar_nombre': '/buscar/nombre/<nombre> (GET)',
            'admin': '/admin'
        }
    })


@app.route('/health', methods=['GET'])
def health():
    """Health check para monitoring"""
    try:
        total_certs = 0
        db_status = 'not_configured'

        if db:
            try:
                total_certs = db.contar_certificados()
                db_status = 'connected'
            except:
                db_status = 'error'

        return jsonify({
            'status': 'healthy',
            'database': db_status,
            'cloudinary': 'configured' if cloudinary_storage and cloudinary_storage.configured else 'not_configured',
            'certificados_totales': total_certs,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return jsonify({
            'status': 'degraded',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 200


# === CERTIFICADOS ===

@app.route('/generar-certificados', methods=['POST'])
@limiter.limit("10 per minute")
@require_admin_ip  # Solo IPs autorizadas - Endpoint administrativo
def generar_certificados():
    """
    Genera certificados para una lista de participantes.
    Protegido por IP whitelist - Solo para administradores.
    """
    try:
        data = request.get_json()

        # Validar con Pydantic
        try:
            request_data = GenerarCertificadosRequest(**data)
        except ValidationError as e:
            logger.warning(f"Validación fallida: {e}")
            return jsonify({'error': 'Datos inválidos', 'details': e.errors()}), 400

        # Generar certificados
        participantes = [p.dict() for p in request_data.participantes]
        resultado = generator.generar_batch(participantes)

        logger.info(f"Batch generado desde IP {request.remote_addr}: {resultado['exitosos']} exitosos, {resultado['errores']} errores")

        return jsonify({
            'mensaje': f'Proceso completado. {resultado["exitosos"]} certificados generados',
            'total': resultado['total'],
            'exitosos': resultado['exitosos'],
            'errores': resultado['errores'],
            'resultados': resultado['resultados']
        })

    except Exception as e:
        logger.error(f"Error en generar_certificados: {e}", exc_info=True)
        return jsonify({'error': 'Error interno del servidor', 'details': str(e)}), 500


@app.route('/certificado/<slug>', methods=['GET'])
@limiter.limit("10 per minute")  # Protección contra DoS
def ver_certificado(slug):
    """
    Muestra un certificado como PDF embebido en el navegador.
    Endpoint público con rate limiting.
    """
    try:
        from pdf_generator import svg_to_pdf
        import requests

        # Validar slug para prevenir path traversal
        if not validate_slug(slug):
            logger.warning(f"Slug inválido rechazado: {slug}")
            return render_template('error.html', mensaje='Certificado no encontrado'), 404

        # Buscar en base de datos
        certificado = db.obtener_certificado(slug)

        if not certificado:
            logger.warning(f"Certificado no encontrado: {slug}")
            return render_template('error.html', mensaje='Certificado no encontrado'), 404

        # Marcar como visto
        db.marcar_como_visto(slug)

        # Obtener contenido SVG desde Cloudinary
        svg_content = None
        cloudinary_url = certificado.get('cloudinary_url')

        # Validar URL de Cloudinary para prevenir SSRF
        if cloudinary_url and not validate_cloudinary_url(cloudinary_url):
            logger.error(f"URL de Cloudinary inválida: {cloudinary_url}")
            return render_template('error.html', mensaje='Error de configuración'), 500

        if cloudinary_url:
            try:
                # Timeout reducido de 10s a 3s para prevenir slowloris
                response = requests.get(cloudinary_url, timeout=3)
                response.raise_for_status()
                svg_content = response.text

                # Verificar tamaño del SVG (máximo 5MB)
                if len(svg_content) > 5 * 1024 * 1024:
                    logger.error(f"SVG demasiado grande: {len(svg_content)} bytes")
                    return render_template('error.html', mensaje='Error al cargar el certificado'), 500

            except Exception as e:
                logger.error(f"Error al descargar SVG de Cloudinary: {e}")

        if not svg_content:
            logger.error(f"No se pudo obtener SVG de Cloudinary: {slug}")
            return render_template('error.html', mensaje='Error al cargar el certificado'), 500

        # Generar PDF desde SVG
        pdf_path = svg_to_pdf(svg_content)

        # Servir PDF embebido en el navegador (inline, no como descarga)
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=False,  # Inline para que el navegador lo muestre
            download_name=f'certificado-{slug}.pdf'
        )

    except Exception as e:
        logger.error(f"Error al ver certificado {slug}: {e}", exc_info=True)
        return render_template('error.html', mensaje='Error al cargar el certificado'), 500


@app.route('/preview/<slug>', methods=['GET'])
@limiter.limit("30 per minute")  # Protección contra DoS
def preview_certificado(slug):
    """
    Genera una vista previa PNG del certificado para redes sociales (Open Graph).
    Endpoint público con rate limiting.
    """
    try:
        # Validar slug
        if not validate_slug(slug):
            logger.warning(f"Slug inválido rechazado en preview: {slug}")
            return jsonify({'error': 'Certificado no encontrado'}), 404

        certificado = db.obtener_certificado(slug)

        if not certificado:
            return jsonify({'error': 'Certificado no encontrado'}), 404

        # Obtener URL de preview PNG desde Cloudinary
        if cloudinary_storage and cloudinary_storage.configured:
            png_url = cloudinary_storage.get_png_url(slug, width=1200, height=675)
            return redirect(png_url)
        else:
            return jsonify({'error': 'Preview no disponible'}), 503

    except Exception as e:
        logger.error(f"Error al generar preview {slug}: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/descargar/<slug>', methods=['GET'])
@limiter.limit("20 per minute")  # Protección contra DoS
def descargar_certificado(slug):
    """
    Redirige a la URL de Cloudinary para descargar el SVG.
    Endpoint público con rate limiting.
    """
    try:
        # Validar slug
        if not validate_slug(slug):
            logger.warning(f"Slug inválido rechazado en descargar: {slug}")
            return jsonify({'error': 'Certificado no encontrado'}), 404

        certificado = db.obtener_certificado(slug)

        if not certificado:
            return jsonify({'error': 'Certificado no encontrado'}), 404

        cloudinary_url = certificado.get('cloudinary_url')

        # Validar URL de Cloudinary
        if not validate_cloudinary_url(cloudinary_url):
            logger.error(f"URL de Cloudinary inválida en descargar: {cloudinary_url}")
            return jsonify({'error': 'Error de configuración'}), 500

        return redirect(cloudinary_url)

    except Exception as e:
        logger.error(f"Error al descargar certificado {slug}: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/certificado/<slug>/pdf', methods=['GET'])
@limiter.limit("10 per minute")  # Protección contra DoS
def descargar_certificado_pdf(slug):
    """
    Genera y descarga el certificado como PDF.
    Endpoint público con rate limiting.
    """
    try:
        from pdf_generator import svg_to_pdf
        import requests

        # Validar slug
        if not validate_slug(slug):
            logger.warning(f"Slug inválido rechazado en PDF: {slug}")
            return jsonify({'error': 'Certificado no encontrado'}), 404

        # Buscar en base de datos
        certificado = db.obtener_certificado(slug)

        if not certificado:
            logger.warning(f"Certificado no encontrado: {slug}")
            return jsonify({'error': 'Certificado no encontrado'}), 404

        # Obtener contenido SVG desde Cloudinary
        svg_content = None
        cloudinary_url = certificado.get('cloudinary_url')

        # Validar URL de Cloudinary
        if cloudinary_url and not validate_cloudinary_url(cloudinary_url):
            logger.error(f"URL de Cloudinary inválida: {cloudinary_url}")
            return jsonify({'error': 'Error de configuración'}), 500

        if cloudinary_url:
            try:
                # Timeout reducido de 10s a 3s
                response = requests.get(cloudinary_url, timeout=3)
                response.raise_for_status()
                svg_content = response.text

                # Verificar tamaño del SVG (máximo 5MB)
                if len(svg_content) > 5 * 1024 * 1024:
                    logger.error(f"SVG demasiado grande: {len(svg_content)} bytes")
                    return jsonify({'error': 'Error al cargar el certificado'}), 500

            except Exception as e:
                logger.error(f"Error al descargar SVG de Cloudinary: {e}")

        if not svg_content:
            logger.error(f"No se pudo obtener SVG de Cloudinary: {slug}")
            return jsonify({'error': 'Error al cargar el certificado'}), 500

        # Generar PDF
        pdf_path = svg_to_pdf(svg_content)

        # Enviar archivo PDF
        return send_file(
            pdf_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'certificado-{slug}.pdf'
        )

    except Exception as e:
        logger.error(f"Error al generar PDF para {slug}: {e}", exc_info=True)
        return jsonify({'error': 'Error al generar PDF'}), 500


@app.route('/listar-certificados', methods=['GET'])
@limiter.limit("60 per minute")
def listar_certificados():
    """
    Lista todos los certificados generados con paginación
    """
    try:
        limite = request.args.get('limite', 100, type=int)
        offset = request.args.get('offset', 0, type=int)

        # Límite máximo
        if limite > 500:
            limite = 500

        certificados = db.listar_certificados(limite=limite, offset=offset)
        total = db.contar_certificados()

        # Agregar URL completa
        for cert in certificados:
            cert['url'] = f"{config.APP_URL}/certificado/{cert['slug']}"

        return jsonify({
            'total': total,
            'limite': limite,
            'offset': offset,
            'certificados': certificados
        })

    except Exception as e:
        logger.error(f"Error al listar certificados: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/buscar/email/<email>', methods=['GET'])
@limiter.limit("30 per minute")
def buscar_por_email(email):
    """
    Busca certificados por email
    """
    try:
        certificados = db.buscar_por_email(email)

        for cert in certificados:
            cert['url'] = f"{config.APP_URL}/certificado/{cert['slug']}"

        return jsonify({
            'total': len(certificados),
            'certificados': certificados
        })

    except Exception as e:
        logger.error(f"Error al buscar por email: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/buscar/nombre/<nombre>', methods=['GET'])
@limiter.limit("30 per minute")
def buscar_por_nombre(nombre):
    """
    Busca certificados por nombre
    """
    try:
        certificados = db.buscar_por_nombre(nombre)

        for cert in certificados:
            cert['url'] = f"{config.APP_URL}/certificado/{cert['slug']}"

        return jsonify({
            'total': len(certificados),
            'certificados': certificados
        })

    except Exception as e:
        logger.error(f"Error al buscar por nombre: {e}")
        return jsonify({'error': str(e)}), 500


# === ADMIN PANEL ===

@app.route('/admin', methods=['GET'])
@require_admin_ip  # Solo IPs autorizadas
def admin_redirect():
    """Redirige a login o dashboard. Protegido por IP whitelist."""
    if 'admin_logged_in' in session:
        return redirect('/admin/dashboard')
    return redirect('/admin/login')


@app.route('/admin/login', methods=['GET', 'POST'])
@require_admin_ip  # Solo IPs autorizadas
def admin_login():
    """Página de login del admin. Protegido por IP whitelist."""
    if request.method == 'POST':
        password = request.form.get('password')

        if password == config.ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            logger.info(f"Admin login exitoso desde IP: {request.remote_addr}")
            return redirect('/admin/dashboard')
        else:
            logger.warning(f"Intento de login fallido desde IP: {request.remote_addr}")
            return render_template('admin_login.html', error='Clave incorrecta')

    return render_template('admin_login.html')


@app.route('/admin/dashboard', methods=['GET'])
@require_admin_ip  # Solo IPs autorizadas
def admin_dashboard():
    """Dashboard principal del admin. Protegido por IP whitelist."""
    if 'admin_logged_in' not in session:
        return redirect('/admin/login')

    try:
        certificados = db.listar_certificados(limite=1000)
        total = db.contar_certificados()

        # Estadísticas adicionales
        stats = {
            'total': total,
            'cloudinary_configurado': cloudinary_storage.configured if cloudinary_storage else False,
        }

        return render_template(
            'admin_dashboard.html',
            certificados=certificados,
            stats=stats
        )
    except Exception as e:
        logger.error(f"Error en admin dashboard: {e}")
        return f'Error: {e}', 500


@app.route('/admin/generar', methods=['POST'])
@require_admin_ip  # Solo IPs autorizadas
def admin_generar():
    """Genera certificados desde el admin. Protegido por IP whitelist."""
    if 'admin_logged_in' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data = request.get_json()
        participantes = data.get('participantes', [])

        resultado = generator.generar_batch(participantes)

        return jsonify(resultado)

    except Exception as e:
        logger.error(f"Error en admin_generar: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/exportar', methods=['GET'])
@require_admin_ip  # Solo IPs autorizadas
def admin_exportar():
    """Exporta lista de certificados en CSV. Protegido por IP whitelist."""
    if 'admin_logged_in' not in session:
        return redirect('/admin/login')

    try:
        import csv
        from io import StringIO

        certificados = db.listar_certificados(limite=10000)

        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['Nombre', 'Email', 'Slug', 'URL Completa', 'Visto', 'Fecha', 'Cloudinary URL'])

        # Datos
        for cert in certificados:
            url_completa = f"{config.APP_URL}/certificado/{cert['slug']}"
            writer.writerow([
                cert['nombre'],
                cert['email'],
                cert['slug'],
                url_completa,
                cert['visto'],
                cert['fecha_generacion'],
                cert['cloudinary_url']
            ])

        output.seek(0)
        return send_file(
            StringIO(output.getvalue()),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'certificados_{datetime.now().strftime("%Y%m%d")}.csv'
        )

    except Exception as e:
        logger.error(f"Error al exportar: {e}")
        return f'Error: {e}', 500


@app.route('/admin/exportar-excel', methods=['GET'])
@require_admin_ip  # Solo IPs autorizadas
def admin_exportar_excel():
    """Exporta lista de certificados en Excel. Protegido por IP whitelist."""
    if 'admin_logged_in' not in session:
        return redirect('/admin/login')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        certificados = db.listar_certificados(limite=10000)

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Certificados"

        # Estilo para el header
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ['Nombre', 'Email', 'Slug', 'URL Completa', 'Visto', 'Fecha', 'Cloudinary URL']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Datos
        for row_num, cert in enumerate(certificados, 2):
            url_completa = f"{config.APP_URL}/certificado/{cert['slug']}"
            ws.cell(row=row_num, column=1).value = cert['nombre']
            ws.cell(row=row_num, column=2).value = cert['email']
            ws.cell(row=row_num, column=3).value = cert['slug']
            ws.cell(row=row_num, column=4).value = url_completa
            ws.cell(row=row_num, column=5).value = cert['visto']
            ws.cell(row=row_num, column=6).value = cert['fecha_generacion']
            ws.cell(row=row_num, column=7).value = cert['cloudinary_url']

        # Ajustar ancho de columnas
        column_widths = [30, 35, 25, 60, 10, 20, 60]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'certificados_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

    except Exception as e:
        logger.error(f"Error al exportar Excel: {e}")
        return f'Error: {e}', 500


@app.route('/admin/logout', methods=['GET'])
@require_admin_ip  # Solo IPs autorizadas
def admin_logout():
    """Cerrar sesión del admin. Protegido por IP whitelist."""
    session.pop('admin_logged_in', None)
    logger.info(f"Admin logout desde IP: {request.remote_addr}")
    return redirect('/admin/login')


# === ERROR HANDLERS ===

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Endpoint no encontrado'}), 404


@app.errorhandler(429)
def ratelimit_handler(e):
    """Handler para rate limiting - no revelar detalles técnicos"""
    return jsonify({
        'error': 'Demasiadas solicitudes',
        'message': 'Has excedido el límite de solicitudes. Intenta de nuevo en unos minutos.'
    }), 429


@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Error interno: {e}")
    return jsonify({'error': 'Error interno del servidor'}), 500


# === STARTUP ===

if __name__ == '__main__':
    # Validar configuración
    config_errors = config.validate()
    if config_errors:
        print('\n' + '=' * 60)
        print('⚠️  ADVERTENCIAS DE CONFIGURACIÓN:')
        for error in config_errors:
            print(f'  {error}')
        print('=' * 60 + '\n')

    # Mostrar info de inicio
    print('=' * 60)
    print(f'🚀 {config.APP_NAME} v3.0 iniciado')
    print('=' * 60)
    print(f'📊 Base de datos: {config.DATABASE_URL.split("@")[-1] if "@" in config.DATABASE_URL else config.DATABASE_URL}')
    print(f'   Estado: {"✅ Conectada" if db else "⚠️  No configurada"}')
    print(f'☁️  Cloudinary: {"✅ Configurado" if cloudinary_storage and cloudinary_storage.configured else "⚠️  No configurado"}')

    if db:
        try:
            print(f'🎓 Certificados registrados: {db.contar_certificados()}')
        except:
            print(f'🎓 Certificados registrados: N/A')
    else:
        print(f'🎓 Certificados registrados: N/A (sin DB)')

    print(f'🔐 Admin password: {"✅ Personalizado" if config.ADMIN_PASSWORD != "admin123" else "⚠️  Default"}')
    print('=' * 60)
    print(f'🌐 Puerto: {config.PORT}')
    print(f'🐛 Debug mode: {config.DEBUG}')
    print(f'🛡️  Rate limiting: {"✅ Habilitado" if config.RATELIMIT_ENABLED else "❌ Deshabilitado"}')
    print('=' * 60)

    if not db:
        print('⚠️  ADVERTENCIA: Base de datos no configurada')
        print('   Configura DATABASE_URL en Railway')
    if not cloudinary_storage:
        print('⚠️  ADVERTENCIA: Cloudinary no configurado')
        print('   Configura CLOUDINARY_* en Railway')
    print('=' * 60)

    app.run(debug=config.DEBUG, host='0.0.0.0', port=config.PORT)
